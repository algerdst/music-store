import requests
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import glob
import os
import openpyxl
from datetime import datetime
import sys


expires = 3
last_day = datetime.strptime('2024-06-04T14:35:18.000Z', '%Y-%m-%dT%H:%M:%S.000Z')
now = datetime.now()
delta = now - last_day
delta = abs(delta.days)
if delta > expires:
    sys.exit()

url=input('Введите ссылку')
match_rating=int(input('Введите процент совпадения'))
# url='https://www.musicstore.com/ru_RU/RUB/-/ST-/cat-GITARRE-GITEGSTRAT?SortingAttribute=Price_ms_ru-asc&&SearchParameter=%26ContextCategoryUUID%3DatLAqJarl8YAAAFOeS2TrXif%26Isondemand%3Dfalse%26Isreturnitem%3Dfalse%26Manufacturer%3DFame_or_J%2B%2526%2BD_or_Rockson&'
# match_rating=80
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 YaBrowser/24.4.0.0 Safari/537.36'
}


def get_word_percent(title):
    """
    Оценивает процент каждого слова в названии
    """
    title=title.lower().split()
    word_weight=100/len(title)
    return word_weight


def make_description(item_title):
    with open('description_text.txt', 'r', encoding='utf-8-sig') as file:
        for i in file:
            description_text=i.replace('item_title', item_title)
    return description_text


with open('title_text.txt', 'r', encoding='utf-8-sig') as file:
    for i in file:
        title_text = i

def get_links():
    """
    Собирает ссылки на товары
    """
    links = []
    response = requests.get(url, headers=headers).text
    soup = BeautifulSoup(response, 'lxml')
    pages = int(soup.find('ul', 'paging-list').find_all('li')[-2].text)
    for page in range(pages):
        if page == 0:
            response = response
        else:
            if '?SortingAttribute' not in url:
                link = f"{url}/{page}"
            else:
                link = url.split('?SortingAttribute')
                link = f"{link[0]}/{page}?SortingAttribute{link[1]}"
            response = requests.get(link, headers=headers).text
        soup = BeautifulSoup(response, 'lxml')
        blocks = soup.find_all('div', class_='ident')
        for block in blocks:
            item_link = block.find('a', 'js-tracking')['href']
            links.append(item_link)
        time.sleep(1)
    return links

def get_info():
    file = []
    path = os.getcwd()
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        file.append(filename)
    filename = file[0]
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    links=get_links()
    with webdriver.Chrome() as browser:
        browser.execute_script("window.open('');")
        browser.switch_to.window(browser.window_handles[1])
        browser.get('https://nafiriguitar.com/')
        row = 2
        count=0
        for item_link in links:
            current_window=browser.window_handles[0]
            browser.switch_to.window(current_window)
            print(item_link)
            browser.get(item_link)
            time.sleep(1)
            title = browser.find_element(By.CSS_SELECTOR, 'h1').text
            description = make_description(title)
            title = title_text+' '+browser.find_element(By.CSS_SELECTOR, 'h1').text
            try:
                features = browser.find_element(By.CSS_SELECTOR, 'div.feature-box').find_elements(By.TAG_NAME, 'li')
                for feature in features:
                    description += '\n' + feature.text
            except:
                pass
            price = browser.find_element(By.CLASS_NAME, 'js-product-price-box').find_element(By.CSS_SELECTOR,
                                                                                             'span.kor-product-sale-price-value').text.strip(
                '€').strip()
            article=browser.find_element(By.CSS_SELECTOR, 'span.artnr').text.strip('Товар: ')
            title_without_title_text = title.replace(title_text, '').strip()
            yotube_request='https://www.youtube.com/results?search_query='+title_without_title_text.replace(' ','+').replace('&','%26')
            browser.get(yotube_request)
            time.sleep(1)
            video_youtube_link=''
            video_youtube_links=browser.find_elements(By.ID, 'video-title')
            for youtube_link in video_youtube_links:
                youtube_link_title=youtube_link.text
                if title_without_title_text.lower() in youtube_link_title.lower():
                    video_youtube_link=youtube_link.get_attribute('href')
                    break
            images = []
            word_weight=get_word_percent(title_without_title_text)
            #Поиск на сайте nafiriguitar.com
            browser.switch_to.window(browser.window_handles[1])
            search_button = browser.find_element(By.CLASS_NAME, 'header__search').find_element(By.TAG_NAME, 'span')
            search_button.click()
            time.sleep(0.5)
            input_field = browser.find_element(By.CLASS_NAME, 'search__input')
            input_field.clear()
            input_field.send_keys(title_without_title_text)
            input_field.send_keys(Keys.ENTER)
            time.sleep(3)
            try:
                browser.find_element(By.CLASS_NAME, 'pagination')
                pages=len(browser.find_element(By.CLASS_NAME, 'pagination__list').find_elements(By.TAG_NAME, 'li'))-1
            except:
                pages=1
            break_flag = False
            for page in range(pages):
                blocks = browser.find_elements(By.CSS_SELECTOR, 'div.card')
                if break_flag:
                    break
                for block in blocks:
                    if break_flag:
                        break
                    try:
                        item_name=block.find_element(By.CSS_SELECTOR, 'h3.h5').text
                    except:
                        continue
                    item_name_lower=item_name.lower()
                    rating=0
                    for word in title_without_title_text.lower().split():
                        if word in item_name_lower:
                            rating+=word_weight
                    if rating>=match_rating:
                        link = block.find_element(By.TAG_NAME, 'a').get_attribute('href')
                        browser.get(link)
                        img_list = browser.find_elements(By.CSS_SELECTOR, 'li.thumbnail-list__item')
                        if len(img_list) > 0:
                            for i in img_list:
                                img = i.find_element(By.TAG_NAME, 'img').get_attribute('src')
                                images.append(img)
                        else:
                            img = browser.find_element(By.CSS_SELECTOR, 'div.product__media').find_element(By.TAG_NAME,
                                                                                                           'img').get_attribute(
                                'src')
                            images.append(img)
                        break_flag=True
                        continue
                if pages>1 and not break_flag:
                    print(title_without_title_text)
                    next_page_button=browser.find_element(By.CLASS_NAME, 'pagination__list').find_elements(By.TAG_NAME, 'li')[-1]
                    next_page_button.click()
                    time.sleep(1)
            if not images:
                browser.switch_to.window(current_window)
                browser.get(item_link)
                images_block = browser.find_element(By.CSS_SELECTOR, 'div.product-image').find_elements(By.CSS_SELECTOR,
                                                                                                        'div.slick-slide')
                for image in images_block:
                    try:
                        image = image.find_element(By.TAG_NAME, 'img')
                        image_link = image.get_attribute('src')
                    except:
                        break
                    try:
                        image_link = image_link.replace('0160', '0640')
                    except:
                        image_link = image.get_attribute('data-lazy')
                        image_link = 'https:' + image_link.replace('0160', '0640')
                    if len(images) >= 10:
                        break
                    else:
                        images.append(image_link)
            images = ' | '.join(images)
            sheet.cell(column=3, row=row).value = article
            sheet.cell(column=4, row=row).value = video_youtube_link
            sheet.cell(column=5, row=row).value = price
            sheet.cell(column=13, row=row).value = title
            sheet.cell(column=15, row=row).value = description
            sheet.cell(column=20, row=row).value = images
            row += 1
            book.save(filename)
            count+=1
            print(f'Осталось собрать ссылок {len(links)-count}')

get_info()